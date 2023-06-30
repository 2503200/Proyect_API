from django.forms import modelform_factory
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from openpyxl.workbook import Workbook

from veterinarios.forms import VeterinarioFormulario
from veterinarios.models import Veterinario

#VeterinarioFormulario = modelform_factory(Veterinario, exclude=[])
# Create your views here.
def agregar_veteri(request):
    pagina = loader.get_template('agregar_veteri.html')
    if request.method == 'GET':
        formulario = VeterinarioFormulario
    elif request.method == 'POST':
        formulario = VeterinarioFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def ver_veteri(request, idVeterinario):
    pagina = loader.get_template('ver_veteri.html')
    #veterinario = Veterinario.objects.get(pk=idVeterinario)
    veterinario = get_object_or_404(Veterinario, pk=idVeterinario)
    mensaje = {'veterinario': veterinario}
    return HttpResponse(pagina.render(mensaje, request))

def editar_veteri(request, idVeterinario):
    pagina = loader.get_template('editar_veteri.html')
    veterinario = get_object_or_404(Veterinario, pk=idVeterinario)
    if request.method == 'GET':
        formulario = VeterinarioFormulario(instance=veterinario)
    elif request.method == 'POST':
        formulario = VeterinarioFormulario(request.POST, instance=veterinario)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    mensaje = {'formulario': formulario}
    return HttpResponse(pagina.render(mensaje, request))

def eliminar_veteri (request, idVeterinario):
    veterinario = get_object_or_404(Veterinario, pk=idVeterinario)
    if veterinario:
        veterinario.delete()
        return redirect('inicio')

def generar_reporte (request):
    #veterinarios = Veterinario.objects.all()
    veterinarios = Veterinario.objects.order_by('apellido')
    wb = Workbook()
    ws = wb.active
    ws['B1'] = 'REPORTE DE VETERINARIOS'
    ws.merge_cells('B1:F1')
    ws['B3'] = 'NOMBRE'
    ws['C3'] = 'APELLIDO'
    ws['D3'] = 'SEXO'
    ws['E3'] = 'EMAIL'
    ws['F3'] = 'NUMERO CELULAR'
    cont = 5
    for veterinario in veterinarios:
        ws.cell(row=cont, column=2).value = veterinario.nombre
        ws.cell(row=cont, column=3).value = veterinario.apellido
        ws.cell(row=cont, column=4).value = veterinario.sexo
        ws.cell(row=cont, column=5).value = veterinario.email
        ws.cell(row=cont, column=6).value = veterinario.numero_celular
        cont = cont + 1
    nom_archivo = "ReporteVeterinariosExcel.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nom_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

    return


